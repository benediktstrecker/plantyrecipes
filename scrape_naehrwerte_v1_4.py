
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scrape_naehrwerte_v1_4.py
- Wie v1.3 (robuste Spaltenerkennung + Schreiben mit openpyxl)
- NEU: Farbiges Markieren aller Zellen, die in diesem Lauf befüllt/aktualisiert wurden
  * Standard: dezentes Hellgrün (C6EFCE), ähnlich "Erfolgs"-Markierung

Aufrufbeispiel:
  py scrape_naehrwerte_v1_4.py --input "C:\plantyrecipes\nutrientsextraktion1.xlsx" --sheet ingridient_nutrients --ing-col-index 0 --link-col-index 1

Optional:
  --dry-run  (nur testen, kein Schreiben)
  --ing-col / --link-col    (Hinweise, falls Header besonders sind)
  --ing-col-index / --link-col-index  (0-basiert, A=0, B=1, …)
"""
import re
import argparse
from typing import Dict, Tuple, Optional, Any
import unicodedata
import shutil
import os

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

URL_RE = re.compile(r'^https?://', re.I)

def normalize_text(s: str) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def parse_numeric_with_unit(text: str):
    if text is None:
        return None, None
    t = normalize_text(text)
    m = re.search(r"([-+]?\d{1,3}(?:[\.\s]\d{3})*(?:[\,\.]\d+)?|\d+[\,\.]\d+|\d+)\s*([µu]g|mg|g|kcal|kJ|%)?", t, re.I)
    if not m:
        return None, None
    num_str = m.group(1)
    unit = m.group(2) or None
    num_str = num_str.replace(".", "").replace(" ", "")
    num_str = num_str.replace(",", ".")
    try:
        val = float(num_str)
    except:
        val = None
    if unit:
        unit = unit.replace("ug", "µg").lower()
    return val, unit

def convert_unit(value, from_u, to_u):
    if value is None:
        return None
    if (from_u or "") == (to_u or ""):
        return value
    if from_u in ["ug"]:
        from_u = "µg"
    if to_u in ["ug"]:
        to_u = "µg"

    mass_units = ["µg", "mg", "g"]
    if from_u in mass_units and to_u in mass_units:
        factor = {"µg": 0.001, "mg": 1.0, "g": 1000.0}
        mg_val = value * factor[from_u]
        return mg_val / factor[to_u]

    if from_u == "kJ" and to_u == "kcal":
        return value / 4.184
    if from_u == "kcal" and to_u == "kJ":
        return value * 4.184
    return None

def extract_all_pairs_from_page(url: str) -> Dict[str, str]:
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    pairs: Dict[str, str] = {}

    for tr in soup.find_all("tr"):
        tds = tr.find_all(["td", "th"])
        if len(tds) >= 2:
            key = normalize_text(tds[0].get_text(" "))
            val = normalize_text(tds[1].get_text(" "))
            if key and val:
                pairs[key] = val

    for dt in soup.find_all("dt"):
        dd = dt.find_next_sibling("dd")
        if dd:
            key = normalize_text(dt.get_text(" "))
            val = normalize_text(dd.get_text(" "))
            if key and val:
                pairs[key] = val

    return pairs

def best_match_value(pairs: Dict[str, str], wanted_title: str) -> Optional[str]:
    if wanted_title in pairs:
        return pairs[wanted_title]
    norm_wanted = normalize_text(wanted_title).lower()
    normalized_pairs = {normalize_text(k).lower(): v for k, v in pairs.items()}
    return normalized_pairs.get(norm_wanted)

def coerce_to_multiindex(df_raw: pd.DataFrame) -> pd.DataFrame:
    if isinstance(df_raw.columns, pd.MultiIndex):
        return df_raw
    if len(df_raw) < 2:
        return df_raw
    titles = list(df_raw.iloc[0].astype(str))
    units = list(df_raw.iloc[1].astype(str))
    tuples = list(zip(titles, units))
    new_cols = pd.MultiIndex.from_tuples(tuples)
    df = df_raw.iloc[2:].copy()
    df.columns = new_cols
    df.reset_index(drop=True, inplace=True)
    return df

def detect_url_and_ing_columns(df: pd.DataFrame):
    url_col = None
    max_urls = -1
    for col in df.columns:
        cnt = 0
        series = df[col]
        for v in series:
            if isinstance(v, str) and URL_RE.match(v.strip()):
                cnt += 1
        if cnt > max_urls:
            max_urls = cnt
            url_col = col

    ing_col = None
    if url_col is not None:
        cols = list(df.columns)
        idx = cols.index(url_col)
        if idx > 0:
            ing_col = cols[idx-1]

    if ing_col is None:
        best_score = -1
        best_col = None
        for col in df.columns:
            if col == url_col:
                continue
            text_like = 0
            nonnull = 0
            series = df[col]
            for v in series:
                if pd.isna(v):
                    continue
                nonnull += 1
                if isinstance(v, str) and not URL_RE.match(v.strip()):
                    text_like += 1
            if text_like > best_score:
                best_score = text_like
                best_col = col
        ing_col = best_col

    return url_col, ing_col

def fill_excel(input_path: str, sheet_name: str,
               ing_hint: Optional[str], link_hint: Optional[str],
               ing_index: Optional[int], link_index: Optional[int],
               dry_run: bool = False,
               highlight_color: str = "C6EFCE") -> None:
    """
    highlight_color: RGB hex ohne '#', z. B. 'C6EFCE' (Hellgrün)
    """
    # 1) DataFrame aufbauen
    try:
        df = pd.read_excel(input_path, sheet_name=sheet_name, header=[0,1])
    except Exception:
        df = pd.read_excel(input_path, sheet_name=sheet_name, header=None)
        df = coerce_to_multiindex(df)
    if not isinstance(df.columns, pd.MultiIndex):
        df = coerce_to_multiindex(df)

    print(f"[DEBUG] Columns detected ({len(df.columns)}):")
    for col in df.columns:
        print("  -", col)

    cols = list(df.columns)

    def _flatten(col):
        if isinstance(col, tuple):
            return " / ".join([normalize_text(str(x)).lower() for x in col if str(x) != "nan"])
        return normalize_text(str(col)).lower()

    ing_col = cols[ing_index] if ing_index is not None else None
    link_col = cols[link_index] if link_index is not None else None

    if ing_col is None and ing_hint:
        target = ing_hint.strip().lower()
        for col in cols:
            if target in _flatten(col):
                ing_col = col
                break
    if link_col is None and link_hint:
        target = link_hint.strip().lower()
        for col in cols:
            if target in _flatten(col):
                link_col = col
                break

    if link_col is None or ing_col is None:
        detect_link_col, detect_ing_col = detect_url_and_ing_columns(df)
        link_col = link_col or detect_link_col
        ing_col = ing_col or detect_ing_col

    if ing_col is None or link_col is None:
        raise ValueError("Konnte Spalten für Zutaten/Link nicht bestimmen. "
                         "Bitte --ing-col-index/--link-col-index angeben (z. B. 0 und 1).")

    print(f"[DEBUG] Using ing_col: { ing_col } | link_col: { link_col }")

    titles = [c[0] for c in df.columns]
    units = [c[1] for c in df.columns]

    # 2) Scrapen & Werte in df schreiben
    written_cells = []  # (excel_row, excel_col) für spätere Hervorhebung
    for ridx in range(len(df)):
        ingr = df.at[ridx, ing_col] if ridx in df.index else None
        link = df.at[ridx, link_col] if ridx in df.index else None
        if pd.isna(ingr) or pd.isna(link):
            continue
        link = str(link).strip()
        if not URL_RE.match(link):
            continue

        print(f"-> Scrape: {ingr} | {link}")
        try:
            pairs = extract_all_pairs_from_page(link)
        except Exception as e:
            print(f"[WARN] Ladefehler bei {ingr}: {e}")
            continue

        for cidx, col in enumerate(df.columns):
            if cidx < 3:
                continue  # A/B/C überspringen
            title = titles[cidx]
            unit_target = units[cidx]
            if title is None or unit_target is None or str(unit_target).strip().lower() == "unit_code":
                continue

            wanted_title = str(title).strip()
            target_unit = str(unit_target).strip()

            raw = best_match_value(pairs, wanted_title)
            if raw is None:
                continue

            val_src, unit_src = parse_numeric_with_unit(raw)
            if unit_src is None or unit_src == "":
                final_val = val_src
            else:
                if target_unit and unit_src and target_unit != unit_src:
                    conv = convert_unit(val_src, unit_src, target_unit)
                    final_val = conv if conv is not None else val_src
                else:
                    final_val = val_src

            if final_val is not None and not pd.isna(final_val):
                df.iat[ridx, cidx] = final_val
                # Excel-Position vormerken (Header sind 2 Zeilen, df ist 0-basiert)
                excel_row = 2 + 1 + ridx  # 1-basiert: Header(2) + 1 + ridx
                excel_col = cidx + 1      # 1-basiert
                written_cells.append((excel_row, excel_col))

    if dry_run:
        print("[DRY-RUN] Kein Schreiben ins Excel.")
        return

    # 3) Zurückschreiben + Highlight
    backup_path = input_path.replace(".xlsx", "_backup.xlsx")
    if os.path.abspath(input_path) != os.path.abspath(backup_path):
        shutil.copyfile(input_path, backup_path)

    wb = load_workbook(input_path)
    ws = wb[sheet_name]

    # Werte schreiben
    for ridx in range(len(df)):
        excel_row = 2 + 1 + ridx
        for cidx in range(len(df.columns)):
            if cidx < 3:
                continue
            val = df.iat[ridx, cidx]
            if pd.isna(val):
                continue
            excel_col = cidx + 1
            ws.cell(row=excel_row, column=excel_col, value=float(val) if isinstance(val, (int, float)) else val)

    # Hervorheben der geschriebenen Zellen
    fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")
    for (r, c) in written_cells:
        ws.cell(row=r, column=c).fill = fill

    wb.save(input_path)
    print(f"[OK] Geschrieben. Backup: {backup_path}. Hervorgehobene Zellen: {len(written_cells)}")
    return

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Pfad zur Master-Excel")
    ap.add_argument("--sheet", default="ingridient_nutrients", help="Sheet-Name")

    ap.add_argument("--ing-col", default=None, help="Teilname/Hint für Zutaten-Spalte (z. B. 'ingridients'/'name')")
    ap.add_argument("--link-col", default=None, help="Teilname/Hint für Link-Spalte (z. B. 'link')")

    ap.add_argument("--ing-col-index", type=int, default=None, help="Index der Zutaten-Spalte (0-basiert)")
    ap.add_argument("--link-col-index", type=int, default=None, help="Index der Link-Spalte (0-basiert)")

    ap.add_argument("--dry-run", action="store_true", help="Kein Schreiben – nur Testlauf.")
    ap.add_argument("--highlight-color", default="C6EFCE", help="Hex-Farbe ohne '#', z. B. 'FFF2CC' für Gelb")
    args = ap.parse_args()

    fill_excel(args.input, args.sheet,
               ing_hint=args.ing_col, link_hint=args.link_col,
               ing_index=args.ing_col_index, link_index=args.link_col_index,
               dry_run=args.dry_run, highlight_color=args.highlight_color)

if __name__ == "__main__":
    main()
